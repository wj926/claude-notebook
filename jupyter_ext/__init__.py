"""Jupyter Notebook 6 server extension for Claude Notebook."""

import hashlib
import io
import json
import os
import re
import time
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from tornado import web
from tornado.ioloop import IOLoop
from notebook.utils import url_path_join as ujoin
from notebook.base.handlers import IPythonHandler

from .jsonio import write_json_atomic

# Max text file size to serve inline (10 MB); larger files get a too-large hint
_MAX_TEXT_PREVIEW = 10 * 1024 * 1024

STATIC_DIR = Path(__file__).parent.parent / "static"

# ---------------------------------------------------------------------------
# Snapshot storage
# ---------------------------------------------------------------------------
# Every save captures the PREVIOUS content to a per-file history directory
# under the user's home, keyed by a hash of the absolute path. This gives
# users a safety net when auto-save is enabled, without polluting the
# workspace with .bak files.

SNAPSHOT_ROOT = Path.home() / ".claude-notebook" / "snapshots"
MAX_SNAPSHOTS_PER_FILE = 20
_SNAPSHOT_TS_RE = re.compile(r'^\d{8}-\d{6}-\d{3}$')


def _snapshot_dir_for(full_path: Path) -> Path:
    """Return the per-file snapshot directory (created on demand by callers)."""
    key = hashlib.sha1(str(full_path.resolve()).encode("utf-8")).hexdigest()
    return SNAPSHOT_ROOT / key


def _take_snapshot(full_path: Path) -> None:
    """Copy the current contents of *full_path* into its snapshot directory.

    No-op if the file doesn't exist yet (first save). Old snapshots beyond
    MAX_SNAPSHOTS_PER_FILE are pruned, keeping the most recent ones. Any
    IO error is swallowed: snapshots are a best-effort safety net, not a
    hard dependency of save.
    """
    if not full_path.is_file():
        return
    try:
        content = full_path.read_bytes()
    except OSError:
        return

    snap_dir = _snapshot_dir_for(full_path)
    try:
        snap_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        return

    index_file = snap_dir / "index.json"
    index = {}
    if index_file.exists():
        try:
            index = json.loads(index_file.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            index = {}
    index["path"] = str(full_path)
    snaps = index.setdefault("snapshots", [])

    ts = datetime.now().strftime("%Y%m%d-%H%M%S-%f")[:-3]
    snap_file = snap_dir / f"{ts}.bak"
    try:
        snap_file.write_bytes(content)
    except OSError:
        return
    snaps.append({"ts": ts, "size": len(content)})

    if len(snaps) > MAX_SNAPSHOTS_PER_FILE:
        excess = snaps[:-MAX_SNAPSHOTS_PER_FILE]
        index["snapshots"] = snaps[-MAX_SNAPSHOTS_PER_FILE:]
        for old in excess:
            old_ts = old.get("ts", "")
            if not _SNAPSHOT_TS_RE.match(old_ts):
                continue
            try:
                (snap_dir / f"{old_ts}.bak").unlink(missing_ok=True)
            except OSError:
                pass

    try:
        index_file.write_text(
            json.dumps(index, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def is_safe_path(workspace: Path, requested_path: str) -> bool:
    try:
        resolved = (workspace / requested_path).resolve()
        ws_str = str(workspace)
        # 워크스페이스 자체 + 그 하위 경로만 허용. prefix 매칭만 하면
        # /home/dami/wj 가 워크스페이스일 때 /home/dami/wj2/... 가
        # startswith 통과하는 escape 회귀 (codex round 5 지적).
        return resolved == workspace or str(resolved).startswith(ws_str + os.sep)
    except (ValueError, OSError):
        return False


def get_git_remote_url(dir_path: Path):
    git_config = dir_path / ".git" / "config"
    if not git_config.is_file():
        return None
    try:
        text = git_config.read_text(encoding="utf-8")
        match = re.search(r'\[remote "origin"\][^\[]*url\s*=\s*(.+)', text)
        if not match:
            return None
        url = match.group(1).strip()
        ssh_match = re.match(r'git@github\.com:(.+?)(?:\.git)?$', url)
        if ssh_match:
            return f"https://github.com/{ssh_match.group(1)}"
        https_match = re.match(r'(https://github\.com/.+?)(?:\.git)?$', url)
        if https_match:
            return https_match.group(1)
        return None
    except (OSError, ValueError):
        return None


SKIP_DIRS = {
    '__pycache__', 'node_modules', '.git', '.venv', 'venv',
    '.Trash', 'Library', '.cache', '.local', '.npm', '.nvm',
    '.zsh_sessions', '.ipython', '.claude',
}


def posix_rel(path: Path, base: Path) -> str:
    """Return path relative to base, always using forward slashes."""
    return path.relative_to(base).as_posix()


def get_directory_listing(dir_path: Path, rel_base: Path) -> list:
    """List a single directory level (non-recursive). Fast."""
    items = []
    try:
        entries = sorted(dir_path.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower()))
    except PermissionError:
        return items
    for entry in entries:
        if entry.name in SKIP_DIRS:
            continue
        rel = posix_rel(entry, rel_base)
        try:
            stat = entry.stat()
            mtime = stat.st_mtime
            size = stat.st_size if entry.is_file() else None
        except OSError:
            mtime = None
            size = None
        if entry.is_dir():
            node = {
                "name": entry.name,
                "path": rel,
                "type": "directory",
                "has_children": True,
                "mtime": mtime,
            }
            repo_url = get_git_remote_url(entry)
            if repo_url:
                node["repo_url"] = repo_url
            items.append(node)
        else:
            items.append({
                "name": entry.name,
                "path": rel,
                "type": "file",
                "mtime": mtime,
                "size": size,
            })
    return items


CONFIG_DIR = Path(__file__).parent.parent / "config"
CONFIG_DIR.mkdir(exist_ok=True)

# Legacy location — migrate on first read
_LEGACY_NAMES_FILE = Path(__file__).parent.parent / ".terminal_names.json"
NAMES_FILE = CONFIG_DIR / "terminal-names.json"


def _read_config(filepath):
    """Read a JSON config file, return dict or empty dict on failure."""
    try:
        return json.loads(filepath.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _write_config(filepath, data):
    """Write a JSON config file atomically."""
    write_json_atomic(filepath, data)


def _read_names():
    """Read terminal config. Migrates from legacy location if needed."""
    # Migrate legacy file
    if _LEGACY_NAMES_FILE.is_file() and not NAMES_FILE.is_file():
        _LEGACY_NAMES_FILE.rename(NAMES_FILE)
    data = _read_config(NAMES_FILE)
    # Migrate legacy format: {name: "displayName"} -> {slot: {display_name, command}}
    if data and all(isinstance(v, str) for v in data.values()):
        migrated = {}
        for i, (k, v) in enumerate(data.items(), 1):
            migrated[str(i)] = {"display_name": v, "command": ""}
        _write_names(migrated)
        return migrated
    return data


def _write_names(data):
    _write_config(NAMES_FILE, data)


TERMINAL_UPLOAD_DIR = "uploads"  # Fixed subdir for terminal uploads


def unique_filepath(dest: Path, fname: str) -> Path:
    """Return a non-colliding path: name.ext -> name (1).ext -> name (2).ext ..."""
    fpath = dest / fname
    if not fpath.exists():
        return fpath
    stem = Path(fname).stem
    suffix = Path(fname).suffix
    i = 1
    while True:
        candidate = dest / f"{stem} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


IMAGE_CONTENT_TYPES = {
    '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
    '.gif': 'image/gif', '.webp': 'image/webp', '.svg': 'image/svg+xml',
    '.ico': 'image/x-icon', '.bmp': 'image/bmp',
}

# Audio / video — served with correct Content-Type so <audio>/<video> tags
# inside rendered markdown play directly. Without these, the default
# application/octet-stream falls back to a download instead.
MEDIA_CONTENT_TYPES = {
    # Audio
    '.mp3': 'audio/mpeg', '.wav': 'audio/wav', '.ogg': 'audio/ogg',
    '.oga': 'audio/ogg', '.m4a': 'audio/mp4', '.aac': 'audio/aac',
    '.flac': 'audio/flac', '.opus': 'audio/ogg',
    # Video
    '.mp4': 'video/mp4', '.m4v': 'video/mp4', '.webm': 'video/webm',
    '.ogv': 'video/ogg', '.mov': 'video/quicktime',
    # Documents — PDFs render in the browser's native viewer when served
    # with this Content-Type via raw=1.
    '.pdf': 'application/pdf',
}


# ---------------------------------------------------------------------------
# Base handler with shared helpers
# ---------------------------------------------------------------------------

class BaseHandler(IPythonHandler):
    """Common helpers shared by all Claude Notebook handlers."""

    def json_response(self, data):
        """Serialize *data* as JSON and finish the response."""
        self.set_header("Content-Type", "application/json; charset=utf-8")
        self.finish(json.dumps(data, ensure_ascii=False))

    def get_xsrf_string(self):
        """Return the XSRF token as a str (decoded from bytes if needed)."""
        xsrf = self.xsrf_token
        if isinstance(xsrf, bytes):
            xsrf = xsrf.decode("utf-8")
        return xsrf

    def inject_script(self, html, **variables):
        """Inject a ``<script>`` block that sets ``window.__KEY = "value"`` for each kwarg."""
        parts = "".join(
            f'window.{key} = {json.dumps(value)};' for key, value in variables.items()
        )
        inject = f'<script>{parts}</script>'
        return html.replace('</head>', inject + '\n</head>')

    def get_workspace(self):
        """Return the workspace root ``Path``."""
        return self.settings["claude_notebook_path"]

    def validate_path(self, file_path):
        """Validate *file_path* against the workspace and return the resolved ``Path``.

        Raises ``tornado.web.HTTPError`` on invalid or missing paths.
        """
        workspace = self.get_workspace()
        if not file_path:
            raise web.HTTPError(400, "path required")
        if not is_safe_path(workspace, file_path):
            raise web.HTTPError(400, "Invalid path: %s" % file_path)
        full_path = (workspace / file_path).resolve()
        return full_path


# ---------------------------------------------------------------------------
# Handlers
# ---------------------------------------------------------------------------

class TerminalNamesHandler(BaseHandler):
    @web.authenticated
    def get(self):
        self.json_response(_read_names())

    @web.authenticated
    def put(self):
        body = json.loads(self.request.body)
        slot = body.get("slot")
        display_name = body.get("display_name", "")
        command = body.get("command", "")
        if not slot:
            raise web.HTTPError(400, "slot required")
        names = _read_names()
        chat_mode = body.get("chat_mode", False)
        names[str(slot)] = {"display_name": display_name, "command": command, "chat_mode": chat_mode}
        _write_names(names)
        self.json_response({"ok": True})

    @web.authenticated
    def delete(self):
        slot = self.get_argument("slot", None)
        if not slot:
            raise web.HTTPError(400, "slot required")
        names = _read_names()
        names.pop(str(slot), None)
        # Re-index slots to keep them sequential
        reindexed = {}
        for i, key in enumerate(sorted(names.keys(), key=lambda x: int(x)), 1):
            reindexed[str(i)] = names[key]
        _write_names(reindexed)
        self.json_response({"ok": True})


class ConfigHandler(BaseHandler):
    """Generic key-value config store. Files saved as config/<key>.json."""

    _VALID_KEY = re.compile(r'^[a-zA-Z0-9_-]+$')

    def _config_path(self, key):
        if not key or not self._VALID_KEY.match(key):
            raise web.HTTPError(400, "Invalid config key: %s" % key)
        return CONFIG_DIR / f"{key}.json"

    @web.authenticated
    def get(self):
        key = self.get_argument("key", None)
        if not key:
            raise web.HTTPError(400, "key required")
        self.json_response(_read_config(self._config_path(key)))

    @web.authenticated
    def put(self):
        body = json.loads(self.request.body)
        key = body.get("key")
        data = body.get("data")
        if not key:
            raise web.HTTPError(400, "key required")
        if data is None:
            raise web.HTTPError(400, "data required")
        _write_config(self._config_path(key), data)
        self.json_response({"ok": True})


class WorkspaceViewerHandler(BaseHandler):
    @web.authenticated
    def get(self):
        base_url = self.settings.get("base_url", "/").rstrip("/")
        viewer_base = base_url + "/claude-notebook"
        xsrf = self.get_xsrf_string()

        focus = self.get_argument("focus", None)
        if focus is None and "/files" in self.request.path:
            focus = "files"

        html = STATIC_DIR.joinpath("index.html").read_text(encoding="utf-8")
        # Insert <base> so relative href/src resolve under /claude-notebook/
        html = html.replace("<head>", f'<head>\n<base href="{viewer_base}/">', 1)
        html = self.inject_script(
            html,
            __VIEWER_BASE=viewer_base,
            __JUPYTER_BASE=base_url,
            __XSRF_TOKEN=xsrf,
            __FOCUS=focus or "",
        )
        self.set_header("Content-Type", "text/html; charset=utf-8")
        self.set_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.finish(html)


class WorkspaceTerminalHandler(BaseHandler):
    @web.authenticated
    def get(self):
        base_url = self.settings.get("base_url", "/").rstrip("/")
        viewer_base = base_url + "/claude-notebook"
        xsrf = self.get_xsrf_string()

        focus = self.get_argument("focus", None)
        if focus is None and "/terminal" in self.request.path:
            focus = "terminal"

        html = STATIC_DIR.joinpath("index.html").read_text(encoding="utf-8")
        html = html.replace("<head>", f'<head>\n<base href="{viewer_base}/">', 1)
        html = self.inject_script(
            html,
            __VIEWER_BASE=viewer_base,
            __JUPYTER_BASE=base_url,
            __XSRF_TOKEN=xsrf,
            __FOCUS=focus or "",
        )
        self.set_header("Content-Type", "text/html; charset=utf-8")
        self.set_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.finish(html)


class LegacyTerminalHandler(BaseHandler):
    @web.authenticated
    def get(self):
        base_url = self.settings.get("base_url", "/").rstrip("/")
        viewer_base = base_url + "/claude-notebook"
        xsrf = self.get_xsrf_string()
        html = STATIC_DIR.joinpath("legacy/terminal.html").read_text(encoding="utf-8")
        # Path replace — copied from original WorkspaceTerminalHandler (lines 412–414)
        html = html.replace('href="terminal.css"',     f'href="{viewer_base}/static/terminal.css"')
        html = html.replace('src="keyboard-guard.js"', f'src="{viewer_base}/static/keyboard-guard.js"')
        html = html.replace('src="terminal.js"',       f'src="{viewer_base}/static/terminal.js"')
        html = self.inject_script(
            html,
            __VIEWER_BASE=viewer_base,
            __JUPYTER_BASE=base_url,
            __XSRF_TOKEN=xsrf,
        )
        self.set_header("Content-Type", "text/html; charset=utf-8")
        self.set_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.finish(html)


class LegacyFilesHandler(BaseHandler):
    @web.authenticated
    def get(self):
        base_url = self.settings.get("base_url", "/").rstrip("/")
        viewer_base = base_url + "/claude-notebook"
        xsrf = self.get_xsrf_string()
        html = STATIC_DIR.joinpath("legacy/index.html").read_text(encoding="utf-8")
        # Path replace — copied from original WorkspaceViewerHandler (lines 393–395)
        html = html.replace('href="/style.css"',         f'href="{viewer_base}/static/legacy/style.css"')
        html = html.replace('src="/keyboard-guard.js"',  f'src="{viewer_base}/static/legacy/keyboard-guard.js"')
        html = html.replace('src="/app.js"',             f'src="{viewer_base}/static/legacy/app.js"')
        # iframe 안에서는 외곽 페이지에 사이드바가 이미 있으므로 옛 사이드바와
        # ☰ 토글을 숨겨서 본문 가로 폭 확보 (이중 사이드바 회귀 방지).
        hide_chrome_css = (
            "<style>"
            ".sidebar, .sidebar-toggle, .sidebar-overlay { display: none !important; }"
            "</style>"
        )
        html = html.replace("</head>", f"{hide_chrome_css}\n</head>", 1)
        html = self.inject_script(
            html,
            __VIEWER_BASE=viewer_base,
            __JUPYTER_BASE=base_url,
            __XSRF_TOKEN=xsrf,
        )
        self.set_header("Content-Type", "text/html; charset=utf-8")
        self.set_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.finish(html)


class WorkspaceStaticHandler(BaseHandler):
    @web.authenticated
    def get(self, filename):
        filepath = STATIC_DIR / filename
        if not filepath.is_file() or not str(filepath.resolve()).startswith(str(STATIC_DIR.resolve())):
            raise web.HTTPError(404)
        content_types = {
            '.css': 'text/css; charset=utf-8',
            '.js': 'application/javascript; charset=utf-8',
            '.html': 'text/html; charset=utf-8',
        }
        ct = content_types.get(filepath.suffix, 'application/octet-stream')
        self.set_header("Content-Type", ct)
        # JS/CSS/HTML iterate fast; tell browsers to revalidate every load
        # so users don't get stuck on a cached pre-fix version after a deploy.
        if filepath.suffix in ('.css', '.js', '.html'):
            self.set_header("Cache-Control", "no-cache, must-revalidate")
        self.finish(filepath.read_bytes())


class WorkspaceTreeHandler(BaseHandler):
    @web.authenticated
    def get(self):
        workspace = self.get_workspace()
        sub = self.get_argument("path", "")
        if sub and not is_safe_path(workspace, sub):
            raise web.HTTPError(400, "Invalid path")
        target = (workspace / sub).resolve() if sub else workspace
        if not target.is_dir():
            raise web.HTTPError(404, "Directory not found")
        tree = get_directory_listing(target, workspace)
        self.json_response(tree)


class WorkspaceFileHandler(BaseHandler):
    @web.authenticated
    async def get(self):
        workspace = self.get_workspace()
        file_path = self.get_argument("path", None)
        raw_mode = self.get_argument("raw", None)
        full_path = self.validate_path(file_path)
        if not full_path.is_file():
            raise web.HTTPError(404, "Not found: %s" % file_path)
        ext = full_path.suffix.lower()
        file_size = full_path.stat().st_size

        # Image, audio/video, or explicit raw mode: stream as binary with the
        # right Content-Type so <img>/<audio>/<video> can render/play.
        if (
            raw_mode is not None
            or ext in IMAGE_CONTENT_TYPES
            or ext in MEDIA_CONTENT_TYPES
        ):
            # raw=1 으로 텍스트 계열 미리보기 (html, svg) 도 inline 렌더 가능하게
            # 적절한 Content-Type 매핑. 모르는 ext 는 octet-stream → 다운로드.
            text_raw_types = {
                '.html': 'text/html; charset=utf-8',
                '.htm':  'text/html; charset=utf-8',
                '.svg':  'image/svg+xml',
            }
            ct = (
                IMAGE_CONTENT_TYPES.get(ext)
                or MEDIA_CONTENT_TYPES.get(ext)
                or text_raw_types.get(ext)
                or 'application/octet-stream'
            )
            self.set_header("Content-Type", ct)
            self.set_header("Content-Length", str(file_size))
            self.set_header("Cache-Control", "public, max-age=3600")
            # Help native media controls show a scrubbable timeline.
            self.set_header("Accept-Ranges", "bytes")
            await self._stream_file(full_path)
            return

        # Text file: check size limit
        if file_size > _MAX_TEXT_PREVIEW:
            self.json_response({
                "path": file_path,
                "name": full_path.name,
                "content": None,
                "extension": ext,
                "too_large": True,
                "size": file_size,
            })
            return

        try:
            content = full_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            # Binary file — stream it
            self.set_header("Content-Type", "application/octet-stream")
            self.set_header("Content-Length", str(file_size))
            await self._stream_file(full_path)
            return
        except Exception as e:
            raise web.HTTPError(500, str(e))

        self.json_response({
            "path": file_path,
            "name": full_path.name,
            "content": content,
            "extension": ext,
        })

    async def _stream_file(self, full_path):
        """Stream a file in 4 MB chunks."""
        chunk_size = 4 * 1024 * 1024
        with open(full_path, "rb") as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                self.write(chunk)
                await self.flush()
        self.finish()


class WorkspaceRawHandler(BaseHandler):
    """Serve workspace files at directory-based URL so HTML 의 상대 경로
    (link href, img src 등) 가 정상 resolve 됨. /api/file?path=... 형식은
    쿼리 URL 이라 상대 경로가 깨지는 문제 회피.

    URL: /claude-notebook/raw/<sub/path/file.ext>
    """

    @web.authenticated
    async def get(self, sub_path):
        workspace = self.get_workspace()
        if not is_safe_path(workspace, sub_path):
            raise web.HTTPError(400, "Invalid path: %s" % sub_path)
        full_path = (workspace / sub_path).resolve()
        if not full_path.is_file():
            raise web.HTTPError(404, "Not found: %s" % sub_path)
        ext = full_path.suffix.lower()
        text_raw_types = {
            '.html': 'text/html; charset=utf-8',
            '.htm':  'text/html; charset=utf-8',
            '.svg':  'image/svg+xml',
            '.css':  'text/css; charset=utf-8',
            '.js':   'application/javascript; charset=utf-8',
            '.json': 'application/json; charset=utf-8',
            '.txt':  'text/plain; charset=utf-8',
            '.md':   'text/markdown; charset=utf-8',
        }
        ct = (
            IMAGE_CONTENT_TYPES.get(ext)
            or MEDIA_CONTENT_TYPES.get(ext)
            or text_raw_types.get(ext)
            or 'application/octet-stream'
        )
        self.set_header("Content-Type", ct)
        self.set_header("Content-Length", str(full_path.stat().st_size))
        self.set_header("Cache-Control", "no-cache, must-revalidate")
        self.set_header("Accept-Ranges", "bytes")
        with open(full_path, "rb") as f:
            while True:
                chunk = f.read(4 * 1024 * 1024)
                if not chunk:
                    break
                self.write(chunk)
                await self.flush()
        self.finish()


class WorkspaceXlsxHandler(BaseHandler):
    """Read an .xlsx (or .xls) and return every sheet as JSON rows.

    The client renders this with the same CSV-style table the .csv viewer
    uses, plus a tab strip for sheet selection. Read-only and best-effort:
    formulas are evaluated to their cached value via data_only=True; styles
    and merges are dropped (we just want the values for preview).
    """

    @web.authenticated
    def get(self):
        file_path = self.get_argument("path", None)
        full_path = self.validate_path(file_path)
        if not full_path.is_file():
            raise web.HTTPError(404, "Not found: %s" % file_path)
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise web.HTTPError(
                501,
                "openpyxl not installed on the server (pip install openpyxl)",
            )
        try:
            wb = load_workbook(full_path, read_only=True, data_only=True)
        except Exception as e:
            raise web.HTTPError(500, "Failed to read xlsx: %s" % e)

        sheets = []
        try:
            for name in wb.sheetnames:
                ws = wb[name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([
                        "" if v is None else
                        (v.isoformat() if hasattr(v, "isoformat") else str(v))
                        for v in row
                    ])
                # Trim trailing all-empty rows so the preview isn't pages of blanks
                while rows and all(c == "" for c in rows[-1]):
                    rows.pop()
                sheets.append({"name": name, "rows": rows})
        finally:
            wb.close()

        self.json_response({
            "path": file_path,
            "name": full_path.name,
            "extension": full_path.suffix.lower(),
            "sheets": sheets,
        })


class WorkspaceUploadHandler(BaseHandler):
    """Upload files to workspace."""
    @web.authenticated
    def post(self):
        workspace = self.get_workspace()
        target_dir = self.get_argument("dir", "")
        if target_dir and not is_safe_path(workspace, target_dir):
            raise web.HTTPError(400, "Invalid dir")
        dest = (workspace / target_dir).resolve() if target_dir else workspace
        if not dest.is_dir():
            raise web.HTTPError(400, "Target is not a directory")
        uploaded = []
        for field_name, file_list in self.request.files.items():
            for f in file_list:
                raw_name = f["filename"]
                # Support relative paths for folder uploads (e.g. "photos/a.png")
                rel = Path(raw_name)
                # Reject absolute or path-traversal attempts
                if rel.is_absolute() or ".." in rel.parts:
                    continue
                fname = str(rel)
                if not fname or not rel.name:
                    continue
                file_dest = dest / rel.parent
                if not str(file_dest.resolve()).startswith(str(workspace.resolve())):
                    continue  # safety check
                file_dest.mkdir(parents=True, exist_ok=True)
                fpath = unique_filepath(file_dest, rel.name)
                fpath.write_bytes(f["body"])
                uploaded.append(posix_rel(fpath, workspace))
        self.json_response({"uploaded": uploaded})


class WorkspaceDeleteHandler(BaseHandler):
    """Delete a file or empty folder."""
    @web.authenticated
    def delete(self):
        workspace = self.get_workspace()
        file_path = self.get_argument("path", None)
        full_path = self.validate_path(file_path)
        if not full_path.exists():
            raise web.HTTPError(404, "Not found")
        # Safety: don't delete workspace root or .agent
        if full_path == workspace or ".agent" in full_path.parts:
            raise web.HTTPError(403, "Cannot delete this path")
        import shutil
        if full_path.is_dir():
            shutil.rmtree(full_path)
        else:
            full_path.unlink()
        self.json_response({"deleted": file_path})


class WorkspaceSaveHandler(BaseHandler):
    """Save (overwrite) a text file in the workspace."""
    @web.authenticated
    def put(self):
        workspace = self.get_workspace()
        body = json.loads(self.request.body)
        file_path = body.get("path")
        content = body.get("content")
        if content is None:
            raise web.HTTPError(400, "content required")
        full_path = self.validate_path(file_path)
        if not full_path.is_file():
            raise web.HTTPError(404, "Not found: %s" % file_path)
        # Snapshot the previous content before overwriting so the user can
        # recover if auto-save commits something they didn't mean to keep.
        _take_snapshot(full_path)
        try:
            full_path.write_text(content, encoding="utf-8")
        except OSError as e:
            raise web.HTTPError(500, "Write error: %s" % str(e))
        self.json_response({"saved": file_path})


class WorkspaceSnapshotsListHandler(BaseHandler):
    """List snapshots for a given file (newest first)."""
    @web.authenticated
    def get(self):
        file_path = self.get_argument("path", None)
        full_path = self.validate_path(file_path)
        snap_dir = _snapshot_dir_for(full_path)
        index_file = snap_dir / "index.json"
        if not index_file.is_file():
            self.json_response({"snapshots": []})
            return
        try:
            index = json.loads(index_file.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            self.json_response({"snapshots": []})
            return
        snaps = list(reversed(index.get("snapshots", [])))
        self.json_response({"snapshots": snaps})


class WorkspaceSnapshotContentHandler(BaseHandler):
    """Return the content of a specific snapshot."""
    @web.authenticated
    def get(self):
        file_path = self.get_argument("path", None)
        ts = self.get_argument("ts", None)
        if not ts or not _SNAPSHOT_TS_RE.match(ts):
            raise web.HTTPError(400, "Invalid ts")
        full_path = self.validate_path(file_path)
        snap_dir = _snapshot_dir_for(full_path)
        snap_file = snap_dir / f"{ts}.bak"
        # Confine to the computed snapshot dir to reject traversal
        try:
            resolved = snap_file.resolve()
            if not str(resolved).startswith(str(snap_dir.resolve())):
                raise web.HTTPError(400, "Invalid snapshot path")
        except (OSError, ValueError):
            raise web.HTTPError(400, "Invalid snapshot path")
        if not snap_file.is_file():
            raise web.HTTPError(404, "Snapshot not found")
        try:
            content = snap_file.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            raise web.HTTPError(415, "Snapshot is not text")
        except OSError as e:
            raise web.HTTPError(500, "Read error: %s" % str(e))
        self.json_response({"content": content, "ts": ts})


class WorkspaceNewFileHandler(BaseHandler):
    """Create a new empty file in the workspace."""
    @web.authenticated
    def post(self):
        body = json.loads(self.request.body)
        file_path = body.get("path")
        if not file_path:
            raise web.HTTPError(400, "path required")
        full_path = self.validate_path(file_path)
        if full_path.exists():
            raise web.HTTPError(409, "Already exists: %s" % file_path)
        try:
            full_path.parent.mkdir(parents=True, exist_ok=True)
            full_path.touch()
        except OSError as e:
            raise web.HTTPError(500, "File creation error: %s" % str(e))
        self.json_response({"created": file_path})


class WorkspaceMkdirHandler(BaseHandler):
    """Create a new folder in the workspace."""
    @web.authenticated
    def post(self):
        body = json.loads(self.request.body)
        dir_path = body.get("path")
        if not dir_path:
            raise web.HTTPError(400, "path required")
        full_path = self.validate_path(dir_path)
        if full_path.exists():
            raise web.HTTPError(409, "Already exists: %s" % dir_path)
        try:
            full_path.mkdir(parents=True, exist_ok=False)
        except OSError as e:
            raise web.HTTPError(500, "mkdir error: %s" % str(e))
        self.json_response({"created": dir_path})


class WorkspaceDeleteMultiHandler(BaseHandler):
    """Delete multiple files/folders in the workspace."""
    @web.authenticated
    def post(self):
        import shutil
        workspace = self.get_workspace()
        body = json.loads(self.request.body)
        paths = body.get("paths", [])
        if not paths:
            raise web.HTTPError(400, "paths required")
        deleted = []
        errors = []
        for file_path in paths:
            if not file_path or not is_safe_path(workspace, file_path):
                errors.append({"path": file_path, "error": "Invalid path"})
                continue
            full_path = (workspace / file_path).resolve()
            if not full_path.exists():
                errors.append({"path": file_path, "error": "Not found"})
                continue
            if full_path == workspace or ".agent" in full_path.parts:
                errors.append({"path": file_path, "error": "Cannot delete this path"})
                continue
            try:
                if full_path.is_dir():
                    shutil.rmtree(full_path)
                else:
                    full_path.unlink()
                deleted.append(file_path)
            except OSError as e:
                errors.append({"path": file_path, "error": str(e)})
        self.json_response({"deleted": deleted, "errors": errors})


class WorkspaceDownloadMultiHandler(BaseHandler):
    """Download multiple files/folders as a single zip."""
    @web.authenticated
    async def post(self):
        workspace = self.get_workspace()
        body = json.loads(self.request.body)
        paths = body.get("paths", [])
        if not paths:
            raise web.HTTPError(400, "paths required")
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for file_path in paths:
                if not file_path or not is_safe_path(workspace, file_path):
                    continue
                full_path = (workspace / file_path).resolve()
                if not full_path.exists():
                    continue
                if full_path.is_dir():
                    for root, dirs, files in os.walk(full_path):
                        root_path = Path(root)
                        if not files and not dirs:
                            arcname = posix_rel(root_path, workspace) + "/"
                            zf.writestr(arcname, "")
                        for fname in files:
                            fpath = root_path / fname
                            arcname = posix_rel(fpath, workspace)
                            zf.write(fpath, arcname)
                else:
                    arcname = posix_rel(full_path, workspace)
                    zf.write(full_path, arcname)
        data = buf.getvalue()
        self.set_header("Content-Type", "application/zip")
        self.set_header("Content-Disposition",
                        "attachment; filename*=UTF-8''selected-files.zip")
        self.set_header("Content-Length", str(len(data)))
        self.write(data)
        self.finish()


class WorkspaceRenameHandler(BaseHandler):
    """Rename a file or folder in the workspace."""
    @web.authenticated
    def put(self):
        body = json.loads(self.request.body)
        old_path = body.get("old_path")
        new_path = body.get("new_path")
        if not old_path or not new_path:
            raise web.HTTPError(400, "old_path and new_path required")
        full_old = self.validate_path(old_path)
        full_new = self.validate_path(new_path)
        if not full_old.exists():
            raise web.HTTPError(404, "Not found: %s" % old_path)
        if full_new.exists():
            raise web.HTTPError(409, "Already exists: %s" % new_path)
        workspace = self.get_workspace()
        if full_old == workspace or ".agent" in full_old.parts:
            raise web.HTTPError(403, "Cannot rename this path")
        try:
            full_old.rename(full_new)
        except OSError as e:
            raise web.HTTPError(500, "Rename error: %s" % str(e))
        self.json_response({"old": old_path, "new": new_path})


class WorkspaceDownloadHandler(BaseHandler):
    """Download a file or folder (as zip) as attachment."""
    @web.authenticated
    async def get(self):
        workspace = self.get_workspace()
        file_path = self.get_argument("path", None)
        if not file_path or not is_safe_path(workspace, file_path):
            raise web.HTTPError(400, "Invalid path")
        full_path = (workspace / file_path).resolve()
        if not full_path.exists():
            raise web.HTTPError(404, "Not found")

        if full_path.is_dir():
            # Folder download: stream as zip
            zip_name = full_path.name + ".zip"
            self.set_header("Content-Type", "application/zip")
            self.set_header("Content-Disposition",
                            f"attachment; filename*=UTF-8''{quote(zip_name)}")
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(full_path):
                    root_path = Path(root)
                    # Add empty directories
                    if not files and not dirs:
                        arcname = posix_rel(root_path, full_path.parent) + "/"
                        zf.writestr(arcname, "")
                    for fname in files:
                        fpath = root_path / fname
                        arcname = posix_rel(fpath, full_path.parent)
                        zf.write(fpath, arcname)
            data = buf.getvalue()
            self.set_header("Content-Length", str(len(data)))
            self.write(data)
            self.finish()
        else:
            # File download: streaming
            file_size = full_path.stat().st_size
            self.set_header("Content-Type", "application/octet-stream")
            self.set_header("Content-Disposition",
                            f"attachment; filename*=UTF-8''{quote(full_path.name)}")
            self.set_header("Content-Length", str(file_size))
            chunk_size = 4 * 1024 * 1024  # 4 MB
            with open(full_path, "rb") as f:
                while True:
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    self.write(chunk)
                    await self.flush()
            self.finish()


# ---------------------------------------------------------------------------
# Chunked upload (large file support)
# ---------------------------------------------------------------------------

# In-memory registry of active chunked uploads: upload_id -> {path, fd, received, total, created}
_chunked_uploads = {}
_CHUNKED_UPLOAD_TTL = 600  # 10 minutes


def _cleanup_stale_uploads():
    """Close and remove chunked uploads older than TTL."""
    now = time.time()
    stale = [uid for uid, e in _chunked_uploads.items() if now - e["created"] > _CHUNKED_UPLOAD_TTL]
    for uid in stale:
        entry = _chunked_uploads.pop(uid)
        try:
            entry["fd"].close()
        except Exception:
            pass
        # Remove incomplete file
        try:
            entry["path"].unlink(missing_ok=True)
        except Exception:
            pass


class ChunkedUploadHandler(BaseHandler):
    """Chunked file upload: init -> append (repeat) -> finalize."""

    @web.authenticated
    def post(self):
        """Initialize a chunked upload. Returns an upload_id."""
        _cleanup_stale_uploads()
        workspace = self.get_workspace()
        body = json.loads(self.request.body)
        filename = body.get("filename", "")
        target_dir = body.get("dir", "")
        total_size = body.get("size", 0)

        if not filename:
            raise web.HTTPError(400, "filename required")

        rel = Path(filename)
        if rel.is_absolute() or ".." in rel.parts:
            raise web.HTTPError(400, "Invalid filename")

        dest = (workspace / target_dir).resolve() if target_dir else workspace
        if not str(dest).startswith(str(workspace)):
            raise web.HTTPError(400, "Invalid dir")
        file_dest = dest / rel.parent
        file_dest.mkdir(parents=True, exist_ok=True)

        fpath = unique_filepath(file_dest, rel.name)
        upload_id = uuid.uuid4().hex

        fd = open(fpath, "wb")
        _chunked_uploads[upload_id] = {
            "path": fpath, "fd": fd, "received": 0,
            "total": total_size, "created": time.time(),
        }

        self.json_response({"upload_id": upload_id, "path": posix_rel(fpath, workspace)})

    @web.authenticated
    def put(self):
        """Append a chunk to an active upload."""
        upload_id = self.get_argument("id", None)
        if not upload_id or upload_id not in _chunked_uploads:
            raise web.HTTPError(400, "Invalid upload_id")

        entry = _chunked_uploads[upload_id]
        chunk = self.request.body
        entry["fd"].write(chunk)
        entry["received"] += len(chunk)

        self.json_response({"received": entry["received"], "total": entry["total"]})

    @web.authenticated
    def delete(self):
        """Finalize (close) or cancel a chunked upload."""
        upload_id = self.get_argument("id", None)
        cancel = self.get_argument("cancel", None)
        if not upload_id or upload_id not in _chunked_uploads:
            raise web.HTTPError(400, "Invalid upload_id")

        entry = _chunked_uploads.pop(upload_id)
        entry["fd"].close()
        if cancel:
            try:
                entry["path"].unlink(missing_ok=True)
            except Exception:
                pass
            self.json_response({"cancelled": True})
        else:
            self.json_response({"path": str(entry["path"]), "size": entry["received"]})


class TerminalUploadHandler(BaseHandler):
    """Upload file to fixed dir for terminal use, return metadata."""
    @web.authenticated
    def post(self):
        workspace = self.get_workspace()
        upload_dir = workspace / TERMINAL_UPLOAD_DIR
        upload_dir.mkdir(exist_ok=True)
        results = []
        for field_name, file_list in self.request.files.items():
            for f in file_list:
                fname = Path(f["filename"]).name
                if not fname:
                    continue
                fpath = unique_filepath(upload_dir, fname)
                fpath.write_bytes(f["body"])
                results.append({
                    "name": fpath.name,
                    "path": str(fpath),
                    "relative": posix_rel(fpath, workspace),
                    "size": len(f["body"]),
                    "content_type": f.get("content_type", "application/octet-stream"),
                })
        self.json_response({"files": results})


# ---------------------------------------------------------------------------
# Extension entry points
# ---------------------------------------------------------------------------

def _jupyter_server_extension_paths():
    return [{"module": "jupyter_ext"}]


def _cleanup_uploads(workspace, max_age_seconds, log):
    """uploads/ 안에서 mtime 이 max_age 지난 파일 삭제. 디렉토리 없으면 no-op."""
    upload_dir = workspace / TERMINAL_UPLOAD_DIR
    if not upload_dir.is_dir():
        return
    cutoff = time.time() - max_age_seconds
    removed = 0
    freed = 0
    for entry in upload_dir.iterdir():
        try:
            # 심볼릭 링크 자체는 보존 (codex 권장 — 사용자가 의도적으로
            # 만든 link 일 수 있음. lstat 으로 link 자신의 mtime 만 보고
            # 타깃은 안 따라감)
            if entry.is_symlink() or not entry.is_file():
                continue
            st = entry.stat()
            if st.st_mtime < cutoff:
                size = st.st_size
                entry.unlink()
                removed += 1
                freed += size
        except OSError as e:
            log.warning("uploads cleanup: skip %s (%s)", entry, e)
    if removed:
        log.info("uploads cleanup: removed %d files (%.1f KB)", removed, freed / 1024)


def _schedule_uploads_cleanup(nb_app):
    """부팅 시 1회 + 1시간마다 uploads/ 7일 TTL 정리.

    Extension 재로딩 시 중복 스케줄러 누적 방지 — nb_app 에 핸들 저장 후
    재진입 시 기존 PeriodicCallback 정지 (codex 권장).
    """
    from tornado.ioloop import IOLoop, PeriodicCallback
    workspace = Path(nb_app.web_app.settings["claude_notebook_path"])
    max_age = 7 * 24 * 3600  # 7일
    interval_ms = 60 * 60 * 1000  # 1시간

    # 중복 가드 — 이미 등록돼있으면 정지
    prev = getattr(nb_app, "_cn_uploads_pcb", None)
    if prev is not None:
        try: prev.stop()
        except Exception: pass

    def _run():
        try:
            _cleanup_uploads(workspace, max_age, nb_app.log)
        except Exception as e:
            nb_app.log.warning("uploads cleanup failed: %s", e)

    IOLoop.current().add_callback(_run)
    pcb = PeriodicCallback(_run, interval_ms)
    pcb.start()
    nb_app._cn_uploads_pcb = pcb


def _auto_create_terminals(nb_app):
    """Create terminals from saved config on server startup."""
    from tornado.ioloop import IOLoop
    from .terminals import set_term_host, set_pending, sync_term_hosts, _write_ssh_command, PTY_READY_DELAY_SEC
    from .hosts import get_host

    def _create():
        term_mgr = nb_app.web_app.settings.get('terminal_manager')
        if term_mgr is None:
            return
        # term-hosts <-> Jupyter live terminals sync
        sync_term_hosts(term_mgr)

        saved = _read_names()
        if not saved:
            return

        # Defensive parse — skip non-numeric slot keys
        valid_slots = []
        for k, v in saved.items():
            try:
                int(k)
            except (TypeError, ValueError):
                nb_app.log.warning("Skip non-numeric slot key: %r", k)
                continue
            valid_slots.append((int(k), k, v))

        migration_dirty = False
        for _, slot, cfg in sorted(valid_slots, key=lambda x: x[0]):
            # host_id 마이그레이션 — 없으면 'local' 채우고 영속화 표시
            if "host_id" not in cfg:
                cfg["host_id"] = "local"
                migration_dirty = True
            host_id = cfg["host_id"]
            host = get_host(host_id) or get_host("local")
            try:
                model = term_mgr.create()
                name = model["name"]
                set_term_host(name, host_id)

                # remote 호스트면 먼저 ssh 진입
                if host and host.get("connect"):
                    IOLoop.current().call_later(
                        PTY_READY_DELAY_SEC,
                        _write_ssh_command, term_mgr, name, host["connect"],
                    )

                # startup command 처리
                command = cfg.get("command", "")
                if command:
                    if host_id == "local":
                        # 기존 동작 — 자동 입력
                        term = term_mgr.get_terminal(name)
                        lines = [l for l in command.split("\n") if l.strip()]
                        for idx, line in enumerate(lines):
                            def _send_line(t=term, cmd=line):
                                try:
                                    t.ptyproc.write(cmd + "\r")
                                except Exception:
                                    pass
                            IOLoop.current().call_later(1.5 + idx * 3, _send_line)
                    else:
                        # remote — 보류 (사용자 클릭 시 실행)
                        set_pending(name, command)

                nb_app.log.info("Auto-created terminal %s (slot %s: %s host=%s)",
                                name, slot, cfg.get("display_name", ""), host_id)
            except Exception as e:
                nb_app.log.warning("Failed to auto-create terminal for slot %s: %s", slot, e)

        if migration_dirty:
            _write_names(saved)

    IOLoop.current().call_later(2, _create)


def load_jupyter_server_extension(nb_app):
    workspace = Path(nb_app.notebook_dir).resolve()
    nb_app.web_app.settings["claude_notebook_path"] = workspace

    # Allow large file uploads — chunk size is 50 MB, set limit with headroom
    _max_body = 100 * 1024 * 1024  # 100 MB (enough for 50 MB chunks + overhead)
    # Patch Tornado HTTPServer settings reliably
    nb_app.tornado_settings.update({
        'max_body_size': _max_body,
        'max_buffer_size': _max_body,
    })
    # Also try to patch already-created server (may or may not exist yet)
    server = getattr(nb_app, 'http_server', None)
    if server is not None:
        server.max_body_size = _max_body
        server.max_buffer_size = _max_body

    # Open new terminals in the workspace directory instead of process cwd
    term_mgr = nb_app.web_app.settings.get('terminal_manager')
    if term_mgr is not None:
        term_mgr.term_settings['cwd'] = str(workspace)

    base_url = nb_app.web_app.settings["base_url"]
    handlers = [
        (ujoin(base_url, r"/claude-notebook"), WorkspaceTerminalHandler),
        (ujoin(base_url, r"/claude-notebook/terminal"), WorkspaceTerminalHandler),
        (ujoin(base_url, r"/claude-notebook/files"), WorkspaceViewerHandler),
        (ujoin(base_url, r"/claude-notebook/static/(.+)"), WorkspaceStaticHandler),
        # /raw/<path> — 디렉토리 기반 URL 로 파일 서빙. HTML inline preview 가
        # 상대 경로 (CSS/img/font) 도 정상 resolve 하도록 (api/file 쿼리 URL
        # 은 base 가 잘못 잡혀 깨짐).
        (ujoin(base_url, r"/claude-notebook/raw/(.+)"), WorkspaceRawHandler),
        (ujoin(base_url, r"/claude-notebook/api/tree"), WorkspaceTreeHandler),
        (ujoin(base_url, r"/claude-notebook/api/file"), WorkspaceFileHandler),
        (ujoin(base_url, r"/claude-notebook/api/xlsx"), WorkspaceXlsxHandler),
        (ujoin(base_url, r"/claude-notebook/api/upload"), WorkspaceUploadHandler),
        (ujoin(base_url, r"/claude-notebook/api/upload-chunk"), ChunkedUploadHandler),
        (ujoin(base_url, r"/claude-notebook/api/save"), WorkspaceSaveHandler),
        (ujoin(base_url, r"/claude-notebook/api/snapshots"), WorkspaceSnapshotsListHandler),
        (ujoin(base_url, r"/claude-notebook/api/snapshots/content"), WorkspaceSnapshotContentHandler),
        (ujoin(base_url, r"/claude-notebook/api/delete"), WorkspaceDeleteHandler),
        (ujoin(base_url, r"/claude-notebook/api/newfile"), WorkspaceNewFileHandler),
        (ujoin(base_url, r"/claude-notebook/api/mkdir"), WorkspaceMkdirHandler),
        (ujoin(base_url, r"/claude-notebook/api/delete-multi"), WorkspaceDeleteMultiHandler),
        (ujoin(base_url, r"/claude-notebook/api/download-multi"), WorkspaceDownloadMultiHandler),
        (ujoin(base_url, r"/claude-notebook/api/rename"), WorkspaceRenameHandler),
        (ujoin(base_url, r"/claude-notebook/api/download"), WorkspaceDownloadHandler),
        (ujoin(base_url, r"/claude-notebook/api/terminal-upload"), TerminalUploadHandler),
        (ujoin(base_url, r"/claude-notebook/api/terminal-names"), TerminalNamesHandler),
        (ujoin(base_url, r"/claude-notebook/api/config"), ConfigHandler),
    ]
    from .hosts import make_handlers as _make_host_handlers
    _host_h = _make_host_handlers(BaseHandler)
    handlers.extend([
        (ujoin(base_url, r"/claude-notebook/api/hosts"),                 _host_h["HostsListHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/hosts/([^/]+)"),         _host_h["HostItemHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/hosts/([^/]+)/test"),    _host_h["HostTestHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/current_host"),          _host_h["CurrentHostHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/ssh-config"),            _host_h["SshConfigHandler"]),
    ])
    from .terminals import make_handlers as _make_term_handlers, sync_term_hosts as _sync_term_hosts
    _term_h = _make_term_handlers(BaseHandler)
    handlers.extend([
        (ujoin(base_url, r"/claude-notebook/api/terminals/new"),                  _term_h["NewSshTerminalHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/term-hosts"),                     _term_h["TermHostsHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/term-hosts/([^/]+)/cleanup"),     _term_h["TermDeleteHookHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/pending-commands"),               _term_h["PendingListHandler"]),
        (ujoin(base_url, r"/claude-notebook/api/pending-commands/([^/]+)"),       _term_h["PendingItemHandler"]),
    ])
    handlers.extend([
        (ujoin(base_url, r"/claude-notebook/legacy-terminal"), LegacyTerminalHandler),
        (ujoin(base_url, r"/claude-notebook/legacy-files"),    LegacyFilesHandler),
    ])
    nb_app.web_app.add_handlers(".*$", handlers)
    nb_app.log.info("Claude Notebook extension loaded at %s/claude-notebook (workspace: %s)", base_url, workspace)

    # Auto-create saved terminals
    _auto_create_terminals(nb_app)

    # uploads/ 7일 TTL 자동 정리
    _schedule_uploads_cleanup(nb_app)
